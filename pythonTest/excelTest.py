import openpyxl
wb = openpyxl.load_workbook('test.xlsx')
ws1 = wb.get_sheet_by_name('mysheet1')
print(ws1['A1'].value)
print(ws1['A2'].value)
ws1['A2'] = 'python'

wb.save("test.xlsx")

# wb = openpyxl.Workbook()
# ws1 = wb.active
# ws1.title ="mysheet1"
# ws2= wb.create_sheet("mysheet2")
# ws1['A1']='fdasfdsa'
# ws1['A2']=30
# ws1['A3']=40
# ws1['A4']='=sum(A2:A3)'
# ws1.append([10,20,30])
# ws1.append([10,20,30])
# 
# ws2['A1']='asdsad'
# 
# wb.save("test.xlsx")
# print('excel write')